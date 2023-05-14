import pandas as pd #pandas for dataframe(df)
from openpyxl import load_workbook #interaction python and excel file
#https://openpyxl.readthedocs.io/en/latest/api/openpyxl.worksheet.worksheet.html
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
import string

class product:
	def __init__(self):
		self.date = None
		self.input_file = None
		self.output_file = None
		self.sheet_name = None
		self.with_chart = True
		self.with_total = True
		self.colomns = None
		self.values = None
		self.index = None
		self.title_report = None

	def process_daily(self):
		# ================== PART 1 - LOAD DATASET
		df = pd.read_excel(self.input_file)
		df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')

		# Salling Total each Gender & each Product Line
		df = df[df['Date'] == self.date]
		df = df.pivot_table(index=self.index, columns=self.colomns, values=self.values, aggfunc='sum').round().fillna(0)

		# export output (format excel)
		df.to_excel(self.output_file, sheet_name=self.sheet_name, startrow=4)
		print('Save dataframe done...')
		if self.with_chart :
			self.create_chart()
		else:
			return df
	
	def create_chart(self):
		# ================== PART 2 - Chart
		wb = load_workbook(self.output_file)
		wb.active = wb[self.sheet_name]

		min_column = wb.active.min_column
		max_column = wb.active.max_column
		min_row = wb.active.min_row
		max_row = wb.active.max_row

		# Bar chart
		barchart = BarChart()
		data = Reference(wb.active, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
		categories = Reference(wb.active, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

		barchart.add_data(data, titles_from_data=True)
		barchart.set_categories(categories)

		wb.active.add_chart(barchart, 'K5')
		barchart.title = 'Sales by Product'
		barchart.style = 2
		wb.save(self.output_file)
		print('Save bar chart to excel...')

		if self.with_total :
			self.create_total(wb,min_column,max_column,min_row,max_row)
		else:
			return 'Done'

	def create_total(self,wb,min_column,max_column,min_row,max_row):
		#Total
		alphabet = list(string.ascii_uppercase)
		alphabet_excel = alphabet[:max_column]

		#[A,B,C,D,E,F,G]
		for i in alphabet_excel:
			if i != 'A':
				wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
				wb.active[f'{i}{max_row+1}'].style = 'Currency'

		wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

		wb.active['A1'] = self.title
		wb.active['A1'].font = Font('Arial', bold=False, size=18)
		wb.save(self.output_file)
		print('Save total gross to excel...')

		return 'Done'